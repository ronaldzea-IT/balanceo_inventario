import openpyxl, json, os, re
from datetime import datetime

print("Iniciando procesamiento fila a fila...")
BASE = os.path.dirname(os.path.abspath(__file__))
ruta   = os.path.join(BASE, "VentasNetas.xlsx")
salida = os.path.join(BASE, "ventas_cache.json")

HOJAS = ["Profit 2K8", "Profit 9.1"]
COLS_BUSCAR = ["co_art","articulo","sucursal","fecha",
               "unidades_facturadas","unidades_devueltas","venta_neta","vendedor","co_prov"]

def limpiar_prov(v):
    if not v:
        return ""
    s = str(v).strip().upper()
    # Si hay multiples codigos separados por coma, tomar el primero
    if "," in s:
        s = s.split(",")[0].strip()
    return s

resumen  = {}
fecha_min = None
fecha_max = None

wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)

for nombre_hoja in HOJAS:
    if nombre_hoja not in wb.sheetnames:
        print(f"  Hoja no encontrada: {nombre_hoja}")
        continue
    print(f"  Procesando hoja: {nombre_hoja}")
    ws = wb[nombre_hoja]

    idx = {}
    encabezado_leido = False
    filas_procesadas = 0

    for fila in ws.iter_rows(values_only=True):
        if not encabezado_leido:
            for i, cell in enumerate(fila):
                if cell is not None:
                    cl = str(cell).lower().strip()
                    for col in COLS_BUSCAR:
                        if cl == col:
                            idx[col] = i
            encabezado_leido = True
            print(f"    Columnas encontradas: {idx}")
            continue

        try:
            sku   = str(fila[idx["co_art"]]).strip() if fila[idx["co_art"]] is not None else ""
            desc  = str(fila[idx["articulo"]]).split("*")[0].strip() if fila[idx["articulo"]] is not None else ""
            suc   = str(fila[idx["sucursal"]]).strip().upper() if fila[idx["sucursal"]] is not None else ""
            fecha = fila[idx["fecha"]]
            uv    = fila[idx["unidades_facturadas"]] or 0
            dev   = fila[idx["unidades_devueltas"]] or 0
            vn    = fila[idx["venta_neta"]] or 0
            ven   = str(fila[idx["vendedor"]]).strip() if "vendedor" in idx and fila[idx["vendedor"]] is not None else "SIN VENDEDOR"
            prov  = limpiar_prov(fila[idx["co_prov"]]) if "co_prov" in idx else ""

            if not sku or sku == "None" or fecha is None:
                continue
            if not isinstance(fecha, datetime):
                continue

            mes = fecha.strftime("%Y-%m")
            if fecha_min is None or fecha < fecha_min: fecha_min = fecha
            if fecha_max is None or fecha > fecha_max: fecha_max = fecha

            k = (sku, desc, suc, mes, ven, prov)
            if k in resumen:
                resumen[k][0] += uv
                resumen[k][1] += dev
                resumen[k][2] += vn
            else:
                resumen[k] = [uv, dev, vn]

            filas_procesadas += 1
            if filas_procesadas % 10000 == 0:
                print(f"    Filas procesadas: {filas_procesadas}")
        except Exception:
            continue

    print(f"  Hoja lista. Total filas: {filas_procesadas}")

wb.close()

print(f"Construyendo JSON con {len(resumen)} registros...")
registros = []
for (sku, desc, suc, mes, ven, prov), v in resumen.items():
    registros.append({
        "sku": sku, "descripcion": desc, "sucursal": suc, "anio_mes": mes,
        "vendedor": ven, "co_prov": prov,
        "unidades_vendidas": int(v[0]), "devoluciones": int(v[1]),
        "ventas_netas": round(float(v[2]), 2)
    })

periodo = {
    "desde": fecha_min.strftime("%b-%y"),
    "hasta": fecha_max.strftime("%b-%y")
}

with open(salida, "w", encoding="utf-8") as f:
    json.dump({"periodo": periodo, "registros": registros}, f, ensure_ascii=False)

print("LISTO. ventas_cache.json generado.")
